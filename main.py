import numpy as np


#'Переменные для работы, тупо копипаста'
#'Qinj1-2,Pb1-2,Pk-2, Pz1-2,D1-2'

import openpyxl
#а вот теперь подгружаем наконец эксельку
wb = openpyxl.load_workbook(filename = 'C:\Учёба\Снова программирование\BMA.xlsm')
sheet=wb['571 вгп']

val = sheet['A1'].value

b = sheet.cell(row = 1,column = 2).value
print(b)
#к хуям массив многомерку, бьём обычный на каждую переменную

Qing=[]
Pk=[]
Pb=[]
Pz=[]
Kl=[]
D=[]
#введём коэффициенты на коридор
kq=0.1
kb=0.1
kk=0.1
kz=0.1
kl=0.15

j=1

while j<32:
    Qing[j] = sheet.cell(j,5).value
    Pb[j] = sheet.cell(j,7).value
    D[j] = sheet.cell(j, 10).value
    Pk[j] = sheet.cell(j, 12).value
    Kl[ j] = sheet.cell(j, 15).value
    Pz[j] = sheet.cell(j, 8).value
    j = j+1

j=2
while j<32:
    if Qing[j]-Qing[j-1]>kq*Qing[j-1] and Pk[j]-Pk[j-1]>kk*Pk[j-1]:
        if Kl[j]-Kl[j-1]>kl*Kl[j-1]:
            print('у нас АвтоГРП или авария, проверить Рзатр срочно!')
            if Pz[j]-Pz[j-1]>kz*Pz[j-1]:
                print('Авария пакера, срочно провести ревизию пакера!')
            elif Pz[j]-Pz[j-1]<=kz*Pz[j-1]:
                print('АвтоГРП, сменить режим работы скважины')

        elif abs(Kl[j]-Kl[j-1])<kl*Kl[j-1]:
            print('Норма, мсена технологического режима')
        elif Kl[j]-Kl[j-1]<(-1)*kl*Kl[j-1]:
            print('Кольматация на смене режима, для уточнения проверить Рбуф')
            if Qing[j]/Pb[j]-Qing[j-1]/Pb[j-1]<(-1)*kl*Qing[j-1]/Pb[j-1]:
                print('Кольматация, требуется обработка ПЗП, пнуть КРСников на это')
            elif Qing[j]/Pb[j]-Qing[j-1]/Pb[j-1]>(-1)*kl*Qing[j-1]/Pb[j-1]:
               print('херня какая-то, пнуть операторов на проверку систем телеметрии')
    if Qing[j] - Qing[j - 1] > kq * Qing[j - 1] and abs(Pk[j] - Pk[j - 1]) < kk * Pk[j - 1]:
        print('Проверить Рбуф, без него уточнение не получится')

  #Пока прервусь и надо полный файл найти, плюс косяк с гитом меня задрал капитально