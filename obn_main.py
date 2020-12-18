import numpy as np
a=3.9+4.4
print(a)
#'Переменные для работы, тупо копипаста'
#'Qinj1-2,Pb1-2,Pk-2, Pz1-2,D1-2'

import openpyxl
#а вот теперь подгружаем наконец эксельку
wb = openpyxl.load_workbook(filename= 'C:\Учёба\Снова программирование\BMA.xlsm')
sheet = wb['571 вгп']

val = sheet['A1'].value
print(val)
b = sheet.cell(row=5, column=10).value
print(b)
#к хуям массив многомерку, бьём обычный на каждую переменную

Qing=[]
Pk=[]
Pb=[]
Pz=[]
Kl=[]
D=[]
#введём коэффициенты на коридор
kq=0.1
kb=0.1
kk=0.1
kz=0.1
kl=0.15

j=0

while j<31:
    Qing.append(sheet.cell(row=5, column=j+9).value)
    Pb.append(sheet.cell(row=7, column=j+9).value)
    D.append(sheet.cell(row=10, column=j+9).value)
    Pk.append(sheet.cell(row=12, column=j+9).value)
    Kl.append(sheet.cell(row=15, column=j+9).value)
    Pz.append(sheet.cell(row=8, column=j+9).value)
    j = j+1

j=1
while j<31:
    if (Qing[j]==0 and Pk[j]!=0) or (Qing[j]!=0 and Pk[j]==0):
        Qing[j]=Qing[j-1]
        Pk[j]=Pk[j-1]
    elif Qing[j]==0 and Pk[j]==0:
        j=32
    if Qing[j] - Qing[j-1] >kq*Qing[j-1] and Pk[j]-Pk[j-1]>kk*Pk[j-1]:
        if Kl[j]-Kl[j-1]>kl*Kl[j-1]:
            print('у нас АвтоГРП или авария, проверить Рзатр срочно!')
            if Pz[j]-Pz[j-1]>kz*Pz[j-1]:
                print('Авария пакера, срочно провести ревизию пакера!')
            elif Pz[j]-Pz[j-1]<=kz*Pz[j-1]:
                print('АвтоГРП, сменить режим работы скважины')

        elif abs(Kl[j]-Kl[j-1])<kl*Kl[j-1]:
            print('Норма, смена технологического режима')
        elif Kl[j]-Kl[j-1]<(-1)*kl*Kl[j-1]:
            print('Кольматация на смене режима, для уточнения проверить Рбуф')
            if Qing[j]/Pb[j]-Qing[j-1]/Pb[j-1]<(-1)*kl*Qing[j-1]/Pb[j-1]:
                print('Кольматация, требуется обработка ПЗП, пнуть КРСников на это')
            elif Qing[j]/Pb[j]-Qing[j-1]/Pb[j-1]>=(-1)*kl*Qing[j-1]/Pb[j-1]:
               print('херня какая-то, пнуть операторов на проверку систем телеметрии')
    if Qing[j] - Qing[j - 1] > kq * Qing[j - 1] and abs(Pk[j] - Pk[j - 1]) < kk * Pk[j - 1]:
        print('Проверить Рбуф, без него уточнение не получится')
        if Qing[j] - Qing[j+1] > Qing[j]*(Pk[j]/Pb[j]-1):
            print('Либо автоГРП, либо авария пакера, проверить Рзатр')
            if Pz[j]-Pz[j-1]>kz*Pz[j-1]:
                print('Авария пакера, КРС к бою, провести ревизию')
            else:
                print('АвтоГРП, сменить режим работы скважины пока не попросили сменить работу')
        elif Qing[j] - Qing[j+1] <= Qing[j]*(Pk[j]/Pb[j]-1):
            print('вполне возможна кольматация, проверить Рбуф и Dшт')
            if D[j]!=D[j-1]:
                print('Штуцирование скважины, ревизируем штуцер силами операторов')
            else:
                if Pb[j]-Pb[j+1]>kb*Pb[j-1]:
                    print('Штуцер размыт, заменить')
                else:
                    print('очевидное-невероятное, пнуть операторов на проверку телеметрии')
    if Qing[j] - Qing[j - 1] > kq * Qing[j - 1] and (Pk[j] - Pk[j - 1]) < (-1)*kk * Pk[j - 1]:
        print('АвтоГРП или авария пакера, уточнить Рзатр')
        if Pz[j] - Pz[j - 1] > kz * Pz[j - 1]:
            print('Авария пакера, КРС к бою, провести ревизию')
        else:
            print('АвтоГРП, кореектируем режим работы скважины')
    if abs(Qing[j] - Qing[j - 1]) < kq * Qing[j - 1] and (Pk[j] - Pk[j - 1]) > kk * Pk[j - 1]:
        print('Вероятнее всего кольматация ПЗП, проверить Рбуф')
        if Qing[j]/Pb[j]-Qing[j-1]/Pb[j+1]>kk*Qing[j-1]/Pb[j-1]:
            print('Кольматация, кинуть валенок в КРС на обработк ПЗП')
        else:
            print('неясные проблемы, возможно телеметрия в отпуске')
    if abs(Qing[j] - Qing[j - 1]) < kq * Qing[j - 1] and (Pk[j] - Pk[j - 1]) <(-1)* kk * Pk[j - 1]:
        print('Возможны АвтоГРП или авария, уточнить Рзатр')
        if Pz[j] - Pz[j - 1] > kz * Pz[j - 1]:
            print('Авария пакера, КРС к бою, провести ревизию')
        else:
            print('АвтоГРП, кореектируем режим работы скважины')
    if (Qing[j] - Qing[j - 1]) <(-1)* kq * Qing[j - 1] and (Pk[j] - Pk[j - 1]) < (-1) * kk * Pk[j - 1]:
        if Kl[j]-Kl[j-1]<(-1)*kk*Kl[j-1]:
            print('Режим слетел, всё нормально')
        else:
            print('Либо кольматация, либо штуцирование, проверить Dшт и Рбуф')
            if D[j]<D[j-1]and Pb[j]-Pb[j-1]<(-1)*kb*Pb[j-1]:
                print('Штуцирование, выставить штуцер (на мороз)')
            elif (D[j]>D[j-1]and Pb[j]-Pb[j-1]<(-1)*kb*Pb[j-1]) or (D[j]<D[j-1]and Pb[j]-Pb[j-1]>kb*Pb[j-1]):
                print('Неопознанная ерунда, пнуть операторов на проверку штуцера и телеметрии')
            elif D[j]!=D[j-1] and Pb[j]-Pb[j-1]<(-1)*kb*Pb[j-1]:
                print('Кольматация, пнуть КРС на обработку ПЗП')
    j=j+1








  #Пока прервусь и надо полный файл найти, плюс косяк с гитом меня задрал капитально